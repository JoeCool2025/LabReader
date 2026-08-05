import argparse
import os
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

def read_query_to_df(db_url, query):
    engine = create_engine(db_url)
    with engine.connect() as conn:
        df = pd.read_sql_query(query, conn)
    return df

def append_df_to_excel(df_new, excel_path, sheet_name='Sheet1', unique_keys=None):
    # If file doesn't exist: write new file with sheet
    if not os.path.exists(excel_path):
        df_new.to_excel(excel_path, sheet_name=sheet_name, index=False)
        print(f"Created {excel_path} with sheet {sheet_name} ({len(df_new)} rows).")
        return

    # File exists. Read existing sheet if present, else create new sheet.
    try:
        # Read only the target sheet; if missing, pandas raises or returns error
        df_existing = pd.read_excel(excel_path, sheet_name=sheet_name)
    except ValueError:
        # sheet not found -> write new sheet by using ExcelWriter with mode='a'
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df_new.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Appended new sheet {sheet_name} to {excel_path} ({len(df_new)} rows).")
        return

    # Combine
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)

    # Optionally deduplicate by unique_keys
    if unique_keys:
        keys = [k.strip() for k in unique_keys.split(',') if k.strip()]
        if keys:
            before = len(df_combined)
            # keep='first' preserves existing rows over new duplicates; change to 'last' if you want newest
            df_combined = df_combined.drop_duplicates(subset=keys, keep='first').reset_index(drop=True)
            after = len(df_combined)
            print(f"Dropped {before-after} duplicate rows based on keys {keys}.")

    # Replace the sheet while preserving other sheets/formatting
    wb = load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
        wb.save(excel_path)  # must save after remove so ExcelWriter can append
    # Write the combined dataframe to the sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Updated sheet {sheet_name} in {excel_path} -> {len(df_combined)} total rows.")

def main(args):
    # Build DB URL. If user passes a raw sqlalchemy URL (starts with sqlite:// or postgres...), use it.
    if args.db.startswith('sqlite:///') or args.db.startswith('postgresql') or args.db.startswith('mysql'):
        db_url = args.db
    else:
        # assume local sqlite file
        db_url = f"sqlite:///{args.db}"

    print("Running query...")
    df_new = read_query_to_df(db_url, args.query)
    if df_new is None or df_new.shape[0] == 0:
        print("Query returned no rows; nothing to append.")
        return

    # Optional: re-order columns or drop large columns before writing
    append_df_to_excel(df_new, args.excel, sheet_name=args.sheet, unique_keys=args.unique_keys)

if __name__ == '__main__':
    p = argparse.ArgumentParser(description="Append DB query results to Excel sheet.")
    p.add_argument('--db', required=True, help="Database path or SQLAlchemy URL (default for plain path: sqlite:///path)")
    p.add_argument('--query', required=True, help="SQL query to select rows to append")
    p.add_argument('--excel', required=True, help="Path to Excel workbook to append to")
    p.add_argument('--sheet', default='Sheet1', help="Target sheet name inside workbook")
    p.add_argument('--unique-keys', default='', help="Comma-separated column names to deduplicate on (optional)")
    args = p.parse_args()
    main(args)